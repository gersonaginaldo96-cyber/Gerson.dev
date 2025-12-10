import customtkinter as ctk
import openpyxl
from openpyxl.utils import get_column_letter
ctk.set_appearance_mode("#5e7c99")
# criacao de interface
speed_modes = ['slow', 'medium', 'fast']
app=ctk.CTk()
app.title('centro humnitario de apoio mahesse')
app.wm_title("maos que acolhem mentes que transormam")
app.geometry('400x460')

#criacao de label 
label_titulo=ctk.CTkLabel(app, text='Maos que acolhem, mentes que transformam', bg_color="#9cbedf", corner_radius=12, text_color="#eaf5f3", font=('times new roman', 16, 'bold'))
label_titulo.grid(row=0, column=0, columnspan=3, pady=20, padx=20, sticky='nsew')
# criacao da label nome
label_nome= ctk.CTkLabel(app, text='Nome do doador:', text_color='black', font=('Arial', 12))
label_nome.grid(row=1, column=0, padx=20, pady=10, sticky='w')
# criacao do campo nome
campo_nome= ctk.CTkEntry(app, placeholder_text='exemplo: João Silva', border_color='gray', border_width=2)
campo_nome.grid(row=1, column=1, padx=20, pady=5, sticky='ew')
  
# criacao da label de tipo de doacao
label_tipo_doacao= ctk.CTkLabel(app, text='Tipo de Doação', text_color='black', font=('Arial', 12))
label_tipo_doacao.grid(row=2, column=0, padx=20, pady=10, sticky='w')

# criacao da label de doacao monetaria
def tipo_doacao_monetaria(escolha):
    tipo = escolha.lower()
    if tipo == 'dinheiro' or tipo == 'monetária':
        label_doacao_monetaria.grid()
        campo_doacao_monetaria.grid()
    elif tipo == 'roupas' or tipo == 'alimentos' or tipo == 'brinquedos' or tipo == 'outros':
        label_doacao_monetaria.grid_remove()
        campo_doacao_monetaria.grid_remove()
    else:
        label_doacao_monetaria.grid_remove()
        campo_doacao_monetaria.grid_remove()
label_doacao_monetaria= ctk.CTkLabel(app, text='Doação Monetária (MZN):', text_color='black', font=('Arial', 12))
label_doacao_monetaria.grid(row=3, column=0, padx=20, pady=10, sticky='w')
# criacao do campo de doacao monetaria
campo_doacao_monetaria= ctk.CTkEntry(app, placeholder_text='Digite o valor da doação em meticais', border_color='gray', border_width=2)
campo_doacao_monetaria.grid(row=3, column=1, padx=20, pady=5, sticky='ew')
# label de feadback
feedback_label = ctk.CTkLabel(app, text='', text_color='black', font=('Arial', 12))
feedback_label.grid(row=5, column=0, columnspan=2, padx=20, pady=10)
# escolher tipos de doacao
lista_tipos_doacao = ['dinheiro', 'alimento', 'roupas']
tipo_var = ctk.StringVar(value='dinheiro')
# funcao para atualizar tipo e exibir/esconder campo monetario
def atualizar_tipo(tipo_escolhido):
    tipo_var.set(tipo_escolhido)
    tipo_doacao_monetaria(tipo_escolhido)
menu_tipo = ctk.CTkOptionMenu(app, values=lista_tipos_doacao, command=atualizar_tipo)
menu_tipo.grid(row=2, column=1, padx=20, pady=5, sticky='ew')

#colocar email no rodape
label_email= ctk.CTkLabel(app, text='gersonaginaldo96@gmail.com', text_color='black', font=('Arial', 10))
label_email.grid(row=7, column=0, columnspan=2, padx=20, pady=10)
# mostrar centros onde deseja que a doacao seja entregue

# lista de centros de mocambique
centros_mocambique = ['1.Centro de Acolhimento Maputo', 
                      '2.Centro de Apoio Beira', 
                      '3.Centro Comunitário Nampula', 
                      '4.Centro de Solidariedade Tete',
                      '5.Centro de Doação Chimoio',
                      '6.Centro de Assistência Pemba',
                      '7.Centro de Suporte Quelimane',
                      '8.Centro de Amparo Lichinga',
                      '9.Centro de Solidariedade Xai-Xai',
                      '10.Centro de Apoio Inhambane',
                      '11.Centro de Doação Nacala',
                      '12.Centro de Assistência Songo',
                      '13.Centro de Suporte Manica',
                      '14.Centro de Amparo Mocuba',
                      '15.Centro de Solidariedade Dondo',
                      '16.Centro de Apoio Vilankulo',
                      '17.Centro de Doação Maxixe',
                      '18.Centro de Assistência Gurúè',
                      '19.Centro de Suporte Angoche',
                      '20.Centro de Amparo Balama',
                      "21.Centro de Solidariedade Montepuez",
                      "22.Centro de Apoio Cuamba"]
# variavel para armazenar a selecao do centro
centro_var = ctk.StringVar(value=centros_mocambique[0])
# funcao para atualizar a variavel quando centro eh selecionado
def atualizar_centro(centro_escolhido):
    print(f'DEBUG: Centro selecionado no callback: {centro_escolhido}')
    centro_var.set(centro_escolhido)
# criar menu suspenso para selecionar centro
label_centro = ctk.CTkLabel(app, text='Centro de Entrega:', text_color='black', font=('Arial', 12))
label_centro.grid(row=4, column=0, padx=20, pady=10, sticky='w')
menu_centro = ctk.CTkOptionMenu(app, values=centros_mocambique, command=atualizar_centro, variable=centro_var)
menu_centro.grid(row=5, column=0, columnspan=2, padx=20, pady=5)
# funcao para salvar doacao
def salvar_doacao():
    nome = campo_nome.get().strip()
    tipo_doacao = tipo_var.get()
    centro_selecionado = centro_var.get()
    valor_doacao = campo_doacao_monetaria.get().strip()
    
    if not nome or not tipo_doacao:
        feedback_label.configure(text='Por favor, preencha todos os campos obrigatórios.', text_color='red')
        return
    
    if tipo_doacao == 'dinheiro':
        if not valor_doacao:
            feedback_label.configure(text='Por favor, insira o valor da doação monetária.', text_color='red')
            return
        try:
            valor_doacao = float(valor_doacao)
        except ValueError:
            feedback_label.configure(text='Por favor, insira um valor numérico válido.', text_color='red')
            return
    else:
        valor_doacao = 'N/A'
    
    with open('doacoes.txt', 'a') as arquivo:
        arquivo.write(f'Nome: {nome}, Tipo de Doação: {tipo_doacao}, Valor da Doação: {valor_doacao}\n')
    
    feedback_label.configure(text='Doação salva com sucesso!', text_color='green')
    print(f'DEBUG: centro_selecionado = {centro_selecionado}')
    salvar_em_excel(nome, tipo_doacao, valor_doacao, centro_selecionado)
# criacao do botao salvar
botao_salvar= ctk.CTkButton(app, text='Salvar Doação', command=salvar_doacao, border_color='gray', border_width=2)
botao_salvar.grid(row=6, column=0, columnspan=2, padx=20, pady=20)
botao_salvar.configure(width=150)
# armazenar dados da doacao em um arquivo de texto
def salvar_em_excel(nome, tipo_doacao, valor_doacao, centro):
    print(f'DEBUG salvar_em_excel: nome={nome}, tipo={tipo_doacao}, valor={valor_doacao}, centro={centro}')
    try:
        wb = openpyxl.load_workbook('maos que acolhem mentes que transformam.xlsx')
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Nome', 'Tipo de Doação', 'Valor', 'Centro'])
    
    ws.append([nome, tipo_doacao, valor_doacao, centro])
    wb.save('maos que acolhem mentes que transformam.xlsx')
# criacao de funcionalidades adicionais
app.mainloop()
https://github.com/gersonaginaldo96-cyber/dia9
